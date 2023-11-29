import random as rand
import openpyxl as xl
import getpass
import os
import subprocess


def intro_intentos():
    for intento in range(1, maxintentos + 1): 
        intento_usuario = int(input(f'Intento {intento}: '))
        if intento_usuario == numero_a_adivinar:
            print('¡Ganaste! El numero era:',numero_a_adivinar,'.' )
            return True
        elif intento_usuario < numero_a_adivinar:
            print('Demasiado bajo. Intenta de nuevo.')
        else:
            print('Demasiado alto. Intenta de nuevo.')
    else:
        print(f'¡Agotaste tus intentos! El número era {numero_a_adivinar}.') 
        return False

def dificultad():
    while True:
        complejidad= int(input('Dificultad: \n1. Facil (20 intentos) \n2. Moderado (12 intentos) \n3. Dificil (5 intentos)\nSelecciona: '))
        if complejidad == 1:
            return 20
        elif complejidad == 2:
            return 12
        elif complejidad == 3:
            return 5
        else:
            print('Opcion no valida, elije una de las dificultades.')

def info_listas():
    nombres_de_jugadores.append(nombre_jugador)
    resultados_de_jugadores.append(resultado)
    if maxintentos == 20:
        dificultad_elegida.append("Fácil")
    elif maxintentos == 12:
        dificultad_elegida.append("Moderado")
    elif maxintentos == 5:
        dificultad_elegida.append("Difícil")
        

nombres_de_jugadores = []
resultados_de_jugadores = []
dificultad_elegida = []


print('Bienvenidos al Juego: Adivinador de números!')

while True:
    menuprincipal= int(input('Modo de Juego: \n1. Solitario \n2. 2 jugadores \n3. Estadisticas \n4. Salir \nSeleeciona:'))
    if 1 <= menuprincipal <= 4:
        if menuprincipal == 1:
            nombre_jugador = str(input('Introduce tu nombre: '))
            maxintentos= dificultad()
            numero_a_adivinar= rand.randint(1,1000)
            print('Comienza el juego, suerte!')
            resultado = intro_intentos()
            info_listas()
            
            print('¿Quieres volver al menú principal?')
            print('1. Sí')
            print('2. No')
            eleccion = int(input())
            if eleccion == 1:
                continue
            else:
                break

        if menuprincipal == 2:
            maxintentos= dificultad()
            while True:
                numero_a_adivinar = int(getpass.getpass(prompt='Jugador 1, elige un numero entre 1 y 1000: '))
                if 1 <= numero_a_adivinar <= 1000:
                    break
                else:
                    print('El numero debe ser entre 1 y 1000.')
            nombre_jugador = str(input('Jugador 2, cual es tu nombre? '))
            print(f'Buena suerte {nombre_jugador}, comienza a adivinar!')
            resultado = intro_intentos()
            info_listas()
            
            print('¿Quieres volver al menú principal?')
            print('1. Sí')
            print('2. No')
            eleccion = int(input())
            if eleccion == 1:
                continue
            else:
                break

        if menuprincipal == 3:
            estadistica = xl.Workbook()
            hoja_estadisticas = estadistica['Sheet']
            hoja_estadisticas.title = 'Estadisticas'
            hoja_estadisticas['A1'].value = 'Jugador'
            hoja_estadisticas['B1'].value = 'Resultado'
            hoja_estadisticas['C1'].value = 'Dificultad'
            for i, nombre in enumerate(nombres_de_jugadores):
                hoja_estadisticas.cell(row=i+2, column=1).value = nombre
                hoja_estadisticas.cell(row=i+2, column=2).value = 'Ganó' if resultados_de_jugadores[i] else 'Perdió'
                hoja_estadisticas.cell(row=i+2, column=3).value = dificultad_elegida[i]
            estadistica.save('Estadísticas.xlsx')
            
            ruta_estadistica = os.path.abspath('Estadísticas.xlsx')
            excel_estadistica = ruta_estadistica
            subprocess.Popen(['start', 'excel', excel_estadistica], shell=True)

        if menuprincipal == 4:
            print('Hasta luego, vuelve pronto.')
        break
    else:
        print('Por favor, seleccione de la opcion 1 a la 4')
